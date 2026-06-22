from datetime import date, time

import pytest
from openpyxl import Workbook

from convert_pilot_xlsx import convert_workbook


def _write_sample_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(2, 5).value = date(2026, 5, 1)
    ws.cell(2, 6).value = date(2026, 5, 2)
    ws.cell(5, 5).value = time(9, 0)
    ws.cell(5, 6).value = time(10, 0)
    ws.cell(6, 5).value = time(10, 20)
    ws.cell(6, 6).value = time(11, 30)

    for col, flow, new_f, old_f, pac, defluor_ml_min, remark in [
        (5, 10.0, 0.8, 0.9, "未投加", 100.0, "除氟剂A"),
        (6, 20.0, None, 1.2, 120.0, 600.0, "除氟剂B，超标"),
    ]:
        ws.cell(9, col).value = flow
        ws.cell(12, col).value = 15.0
        ws.cell(13, col).value = 8.0
        ws.cell(14, col).value = 1500.0
        ws.cell(16, col).value = old_f
        ws.cell(17, col).value = new_f
        ws.cell(21, col).value = "7~8"
        ws.cell(24, col).value = "6~7"
        ws.cell(34, col).value = 0.3
        ws.cell(40, col).value = 2.5
        ws.cell(54, col).value = pac
        ws.cell(60, col).value = defluor_ml_min
        ws.cell(68, col).value = 3.0
        ws.cell(77, col).value = remark

    wb.save(path)


def test_convert_workbook_builds_expanded_and_strict_outputs(tmp_path):
    path = tmp_path / "pilot.xlsx"
    _write_sample_workbook(path)

    expanded, strict, report = convert_workbook(path)

    assert len(expanded) == 2
    assert len(strict) == 1
    assert report["rows"]["expanded"] == 2
    assert report["rows"]["strict"] == 1


def test_convert_workbook_normalizes_dose_and_effluent_fields(tmp_path):
    path = tmp_path / "pilot.xlsx"
    _write_sample_workbook(path)

    expanded, _, _ = convert_workbook(path)
    first = expanded.iloc[0]
    second = expanded.iloc[1]

    assert first["pacl_dose"] == 0
    assert first["defluor_dose"] == pytest.approx(0.6)
    assert first["effluent_f"] == 0.8
    assert first["effluent_f_source"] == "new_meter"
    assert first["pacl_tank_ph"] == 7.5
    assert first["defluor_tank_ph"] == 6.5
    assert first["defluor_agent_type"] == "A"

    assert second["pacl_dose"] == 120.0
    assert second["defluor_dose"] == pytest.approx(1.8)
    assert second["effluent_f"] == 1.2
    assert second["effluent_f_source"] == "old_meter"
    assert second["defluor_agent_type"] == "B"
    assert bool(second["is_over_limit_remark"]) is True
