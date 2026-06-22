"""Convert the pilot-test workbook into canonical training CSV files.

The pilot workbook is laid out sideways: each batch is a date column and each
parameter is a row. The model training pipeline expects the opposite layout,
with one sample per row and canonical field names as columns.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core import config as cfg


ROW_IN_TIME = 5
ROW_OUT_TIME = 6
ROW_FLOW = 9
ROW_INFLUENT_F = 12
ROW_INFLUENT_PH = 13
ROW_CONDUCTIVITY = 14
ROW_EFFLUENT_OLD = 16
ROW_EFFLUENT_NEW = 17
ROW_PACL_TANK_PH = 21
ROW_DEFLUOR_TANK_PH = 24
ROW_RECYCLE_FLOW = 34
ROW_WASTE_FLOW = 40
ROW_MAGNETIC_DOSE_KG = 46
ROW_PACL_CALC_PPM = 52
ROW_PACL_PUMP_ML_MIN = 53
ROW_PACL_DOSE_PPM = 54
ROW_DEFLUOR_PUMP_ML_MIN = 60
ROW_DEFLUOR_ACTIVE_PPM = 61
ROW_DEFLUOR_METER_L_MIN = 62
ROW_PAM_DOSE_PPM = 68
ROW_REMARK = 77

FIRST_BATCH_COL = 5


def _json_default(value: Any):
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return value.isoformat()
    return str(value)


def _cell(ws, row: int, col: int):
    return ws.cell(row, col).value


def _numbers(text: str) -> list[float]:
    return [float(x) for x in re.findall(r"[-+]?\d+(?:\.\d+)?", text)]


def parse_numeric(value: Any, *, empty_text_as_zero: bool = False) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if "数据有误" in text or "未记录" in text:
            return None
        nums = _numbers(text)
        if nums:
            if "~" in text or "～" in text:
                return sum(nums[:2]) / 2 if len(nums) >= 2 else nums[0]
            return nums[0]
        if empty_text_as_zero:
            return 0.0
    return None


def parse_effluent(value: Any) -> tuple[float | None, str]:
    if value is None:
        return None, "missing"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value), "ok"
    text = str(value).strip()
    if not text:
        return None, "missing"

    nums = _numbers(text)
    if not nums:
        return None, "non_numeric"

    if "不准" in text or "故障" in text or "推测" in text:
        flag = "estimated_or_fault"
    elif "~" in text or "～" in text:
        flag = "range_midpoint"
    elif len(nums) > 1:
        flag = "ambiguous_parentheses"
    else:
        flag = "ok"

    if "~" in text or "～" in text:
        return sum(nums[:2]) / 2 if len(nums) >= 2 else nums[0], flag
    return nums[0], flag


def _excel_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value).date()
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _combine_datetime(day: date | None, time_value: Any) -> str | None:
    if day is None or time_value is None or not hasattr(time_value, "hour"):
        return None
    stamp = datetime.combine(day, time_value)
    return stamp.isoformat(timespec="minutes")


def _remark_flags(raw: Any) -> dict[str, Any]:
    text = "" if raw is None else str(raw)
    if "除氟剂B" in text:
        agent = "B"
    elif "除氟剂A" in text:
        agent = "A"
    else:
        agent = None
    return {
        "remark_raw": text or None,
        "defluor_agent_type": agent,
        "is_over_limit_remark": "超标" in text,
        "is_low_f_influent": "低氟" in text,
        "is_magnetic_run": "跑磁" in text,
    }


def _build_record(ws, col: int) -> dict[str, Any]:
    day = _excel_date(_cell(ws, 2, col))
    flow = parse_numeric(_cell(ws, ROW_FLOW, col))
    defluor_pump_ml_min = parse_numeric(_cell(ws, ROW_DEFLUOR_PUMP_ML_MIN, col))
    defluor_dose = None
    if flow and defluor_pump_ml_min is not None:
        defluor_dose = defluor_pump_ml_min / (flow * 1000 / 60)

    effluent_new, new_flag = parse_effluent(_cell(ws, ROW_EFFLUENT_NEW, col))
    effluent_old, old_flag = parse_effluent(_cell(ws, ROW_EFFLUENT_OLD, col))
    if effluent_new is not None:
        effluent = effluent_new
        effluent_source = "new_meter"
        effluent_flag = new_flag
    else:
        effluent = effluent_old
        effluent_source = "old_meter" if effluent_old is not None else None
        effluent_flag = old_flag

    record = {
        "sample_col": col,
        "date": day.isoformat() if day else None,
        "influent_sample_time": _combine_datetime(day, _cell(ws, ROW_IN_TIME, col)),
        "effluent_sample_time": _combine_datetime(day, _cell(ws, ROW_OUT_TIME, col)),
        "timestamp": _combine_datetime(day, _cell(ws, ROW_OUT_TIME, col)),
        "influent_flow": flow,
        "influent_ph": parse_numeric(_cell(ws, ROW_INFLUENT_PH, col)),
        "conductivity": parse_numeric(_cell(ws, ROW_CONDUCTIVITY, col)),
        "influent_f": parse_numeric(_cell(ws, ROW_INFLUENT_F, col)),
        "pacl_dose": parse_numeric(_cell(ws, ROW_PACL_DOSE_PPM, col), empty_text_as_zero=True),
        "defluor_dose": defluor_dose,
        "pacl_tank_ph": parse_numeric(_cell(ws, ROW_PACL_TANK_PH, col)),
        "defluor_tank_ph": parse_numeric(_cell(ws, ROW_DEFLUOR_TANK_PH, col)),
        "recycle_flow": parse_numeric(_cell(ws, ROW_RECYCLE_FLOW, col)),
        "waste_flow": parse_numeric(_cell(ws, ROW_WASTE_FLOW, col)),
        "pam_dose": parse_numeric(_cell(ws, ROW_PAM_DOSE_PPM, col)),
        "effluent_f": effluent,
        "effluent_f_new": effluent_new,
        "effluent_f_old": effluent_old,
        "effluent_f_source": effluent_source,
        "effluent_quality_flag": effluent_flag,
        "pacl_dose_calc_ppm": parse_numeric(_cell(ws, ROW_PACL_CALC_PPM, col), empty_text_as_zero=True),
        "pacl_pump_ml_min": parse_numeric(_cell(ws, ROW_PACL_PUMP_ML_MIN, col)),
        "defluor_pump_ml_min": defluor_pump_ml_min,
        "defluor_active_ppm": parse_numeric(_cell(ws, ROW_DEFLUOR_ACTIVE_PPM, col)),
        "defluor_meter_l_min": parse_numeric(_cell(ws, ROW_DEFLUOR_METER_L_MIN, col)),
        "magnetic_dose_kg": parse_numeric(_cell(ws, ROW_MAGNETIC_DOSE_KG, col)),
    }
    record.update(_remark_flags(_cell(ws, ROW_REMARK, col)))
    return record


def _quality_report(expanded: pd.DataFrame, strict: pd.DataFrame, source: Path) -> dict[str, Any]:
    required = list(cfg.MODEL_INPUT_COLS) + [cfg.TARGET_COL]
    missing_counts = {
        col: int(expanded[col].isna().sum())
        for col in required
        if col in expanded.columns
    }
    source_counts = expanded["effluent_f_source"].fillna("missing").value_counts().to_dict()
    flag_counts = expanded["effluent_quality_flag"].fillna("missing").value_counts().to_dict()
    return {
        "source_file": str(source),
        "rows": {
            "candidate_columns": int(len(expanded)),
            "expanded": int(len(expanded.dropna(subset=[cfg.TARGET_COL]))),
            "strict": int(len(strict)),
        },
        "effluent_f_source_counts": {str(k): int(v) for k, v in source_counts.items()},
        "effluent_quality_flag_counts": {str(k): int(v) for k, v in flag_counts.items()},
        "missing_counts": missing_counts,
        "notes": [
            "expanded keeps all rows with a numeric effluent_f target, using new meter first and old meter as fallback",
            "strict keeps only new-meter rows with ok quality and complete required model fields",
            "pacl_dose uses PAC ppm row; non-numeric dose text such as no-dosing is treated as 0",
            "defluor_dose is converted from dosing pump ml/min to mL/L of 15% stock solution",
            "pilot workbook columns are already batch-paired, so row-based HRT delay should not be applied blindly",
        ],
    }


def convert_workbook(path: str | Path) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    source = Path(path)
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb.active

    records = []
    for col in range(FIRST_BATCH_COL, ws.max_column + 1):
        if _cell(ws, 2, col) is None:
            continue
        records.append(_build_record(ws, col))

    expanded = pd.DataFrame(records)
    expanded = expanded.dropna(subset=[cfg.TARGET_COL]).reset_index(drop=True)

    required = list(cfg.MODEL_INPUT_COLS) + [cfg.TARGET_COL]
    strict_mask = (
        (expanded["effluent_f_source"] == "new_meter")
        & (expanded["effluent_quality_flag"] == "ok")
        & expanded[required].notna().all(axis=1)
    )
    strict = expanded.loc[strict_mask].reset_index(drop=True)
    report = _quality_report(expanded, strict, source)
    return expanded, strict, report


def write_outputs(path: str | Path, output_dir: str | Path = "data/prepared",
                  prefix: str = "pilot") -> dict[str, Any]:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    expanded, strict, report = convert_workbook(path)

    expanded_path = output / f"{prefix}_expanded.csv"
    strict_path = output / f"{prefix}_strict.csv"
    report_path = output / f"{prefix}_quality_report.json"

    expanded.to_csv(expanded_path, index=False, encoding="utf-8-sig")
    strict.to_csv(strict_path, index=False, encoding="utf-8-sig")
    report.update({
        "outputs": {
            "expanded_csv": str(expanded_path),
            "strict_csv": str(strict_path),
            "quality_report_json": str(report_path),
        }
    })
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=_json_default),
        encoding="utf-8",
    )
    return report


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", help="Pilot workbook path")
    parser.add_argument("--output-dir", default="data/prepared")
    parser.add_argument("--prefix", default="pilot")
    args = parser.parse_args()
    report = write_outputs(args.workbook, args.output_dir, args.prefix)
    print(json.dumps(report, ensure_ascii=False, indent=2, default=_json_default))


if __name__ == "__main__":
    main()
